# services/documents/marksheet_template.py
"""
FULL MARK SHEET TEMPLATE - Tanzania NECTA
Complete template for all subjects
No title rows - Clean headers only
"""
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from .base_document import BaseDocument

class MarkSheetTemplate(BaseDocument):
    """
    Complete Mark Sheet Template for class teachers
    
    Features:
    - All subjects in one sheet
    - Gender column included
    - Proper formatting
    - No title rows
    """
    
    def __init__(self, student_list, class_info=None, subjects=None, **kwargs):
        """
        Initialize mark sheet template
        
        Args:
            student_list: List of students
            class_info: Dictionary with class details
            subjects: List of subject names
            **kwargs: Additional parameters
        """
        super().__init__(student_list, document_type='marksheet', **kwargs)
        
        self.class_info = class_info or {}
        self.subjects = subjects or ['Mathematics', 'English', 'Kiswahili', 'Science', 'Geography']
        self.grading_rules = class_info.get('grading_rules', 'CSEE')
        
        # Set metadata
        self.metadata.update({
            'class_name': class_info.get('name', 'Form 4'),
            'stream': class_info.get('stream', ''),
            'subjects': self.subjects,
            'student_count': len(student_list) if student_list else 0,
            'grading_rules': self.grading_rules
        })
    
    def generate(self):
        """Generate the complete mark sheet template"""
        # Prepare student data with GENDER column
        students_df = self._prepare_student_data()
        
        # Add subject columns (EMPTY for teachers to fill)
        for subject in self.subjects:
            students_df[subject] = ''
        
        # Add remarks column
        students_df['Remarks'] = ''
        
        self.processed_data = students_df
        return self
    
    def _prepare_student_data(self):
        """Prepare student data for template with GENDER column"""
        students = self.raw_data
        
        if not students:
            # Create sample data for testing
            students = [
                {
                    'admission_no': 'ADM001', 
                    'student_id': 'STU001', 
                    'full_name': 'JOHN MWAMBA', 
                    'gender': 'M',
                    'class': 'Form 4', 
                    'stream': 'East'
                },
                {
                    'admission_no': 'ADM002', 
                    'student_id': 'STU002', 
                    'full_name': 'SARAH JUMANNE', 
                    'gender': 'F',
                    'class': 'Form 4', 
                    'stream': 'East'
                }
            ]
        
        # Create DataFrame with required columns
        df = pd.DataFrame(students)
        
        # Ensure required columns exist (INCLUDING GENDER)
        required_columns = ['admission_no', 'student_id', 'full_name', 'gender', 'class', 'stream']
        
        for col in required_columns:
            if col not in df.columns:
                if col == 'gender':
                    df[col] = 'M'  # Default to Male
                elif col in ['class', 'stream']:
                    df[col] = 'NOT SET'
                else:
                    df[col] = f'{col.upper()}_MISSING'
        
        # Reorder columns
        column_order = ['admission_no', 'student_id', 'full_name', 'gender', 'class', 'stream']
        df = df[column_order]
        
        return df
    
    def to_excel_bytes(self):
        """Generate Excel template as bytes - NO TITLE ROWS"""
        if self.processed_data is None:
            self.generate()
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main data sheet - headers start at row 1
            self.processed_data.to_excel(
                writer, 
                sheet_name='MARK_SHEET', 
                index=False
            )
            
            workbook = writer.book
            worksheet = writer.sheets['MARK_SHEET']
            
            # Apply formatting
            self._apply_excel_formatting(workbook, worksheet)
            
            # Add instructions sheet
            self._add_instructions_sheet(workbook)
        
        output.seek(0)
        return output
    
    def _apply_excel_formatting(self, workbook, worksheet):
        """Apply formatting to Excel worksheet - NO TITLE ROW"""
        # Header styling
        header_fill = PatternFill(
            start_color="366092",  # Professional blue
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Student info columns styling (read-only)
        info_fill = PatternFill(
            start_color="F2F2F2",  # Light gray
            end_color="F2F2F2",
            fill_type="solid"
        )
        
        # Gender column styling
        gender_fill = PatternFill(
            start_color="E2EFDA",  # Light green
            end_color="E2EFDA",
            fill_type="solid"
        )
        
        # Subject columns styling (editable)
        subject_fill = PatternFill(
            start_color="FFF2CC",  # Light yellow - for editing
            end_color="FFF2CC",
            fill_type="solid"
        )
        
        # Remarks column styling
        remarks_fill = PatternFill(
            start_color="DDEBF7",  # Light blue
            end_color="DDEBF7",
            fill_type="solid"
        )
        
        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply to all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                
                # Header row (row 1)
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Data rows (row 2 onwards)
                elif cell.row > 1:
                    col_letter = get_column_letter(cell.column)
                    
                    # Student info columns (A-C) - read-only
                    if cell.column <= 3:  # Columns A-C
                        cell.fill = info_fill
                        cell.font = Font(color="000000", size=10)
                    
                    # Gender column (D) - read-only
                    elif cell.column == 4:  # Column D (gender)
                        cell.fill = gender_fill
                        cell.font = Font(color="000000", size=10, bold=True)
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Class & Stream columns (E-F) - read-only
                    elif cell.column <= 6:  # Columns E-F
                        cell.fill = info_fill
                        cell.font = Font(color="000000", size=10)
                    
                    # Subject columns (G onwards) - EDITABLE
                    elif cell.column <= 6 + len(self.subjects):  # Subject columns
                        cell.fill = subject_fill
                        cell.font = Font(color="000000", size=10)
                        cell.alignment = Alignment(horizontal="center")
                        cell.number_format = '0'  # Integer format
                    
                    # Remarks column (last column) - editable
                    else:
                        cell.fill = remarks_fill
                        cell.font = Font(color="000000", size=9, italic=True)
        
        # Set column widths
        column_widths = {
            'A': 15,  # Admission No
            'B': 12,  # Student ID
            'C': 25,  # Full Name
            'D': 10,  # Gender (M/F)
            'E': 10,  # Class
            'F': 10,  # Stream
            # Subject columns: 12 width
        }
        
        # Set subject column widths
        for i in range(len(self.subjects)):
            col_letter = get_column_letter(7 + i)  # Start from column G
            column_widths[col_letter] = 12
        
        # Remarks column width
        remarks_col = get_column_letter(7 + len(self.subjects))
        column_widths[remarks_col] = 25
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Freeze header row and first 6 columns (student info)
        worksheet.freeze_panes = "G2"
        
        # Add data validation
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Gender validation (M or F only)
        dv_gender = DataValidation(
            type="list",
            formula1='"M,F"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Gender",
            error="Please enter M or F only"
        )
        dv_gender.add(f'D2:D{worksheet.max_row}')
        worksheet.add_data_validation(dv_gender)
        
        # Marks validation for all subject columns (0-100)
        for i in range(len(self.subjects)):
            col_letter = get_column_letter(7 + i)  # Subject columns start at G
            dv_marks = DataValidation(
                type="decimal",
                operator="between",
                formula1=0,
                formula2=100,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid Marks",
                error="Marks must be between 0 and 100"
            )
            dv_marks.add(f'{col_letter}2:{col_letter}{worksheet.max_row}')
            worksheet.add_data_validation(dv_marks)
    
    def _add_instructions_sheet(self, workbook):
        """Add instructions sheet"""
        instructions_ws = workbook.create_sheet(title="MAELEKEZO")
        
        subjects_list = "\n".join([f"   - {subject}" for subject in self.subjects])
        
        instructions = [
            ["📘 MAELEKEZO YA FOMU YA ALAMA ZOTE"],
            [""],
            ["TAARIFA ZA DARASA:"],
            [f"Darasa: {self.class_info.get('name', 'Form 4')}"],
            [f"Mzunguko: {self.class_info.get('stream', '')}"],
            [f"Mfumo wa upimaji: NECTA {self.grading_rules}"],
            [f"Idadi ya masomo: {len(self.subjects)}"],
            [""],
            ["📝 JINSI YA KUJAZA:"],
            ["1. USIBADILISHE TAARIFA ZA MWANAFUNZI:"],
            ["   - Namba ya usajili (Safu A)"],
            ["   - Kitambulisho (Safu B)"],
            ["   - Jina kamili (Safu C)"],
            ["   - Jinsia (Safu D) - M au F tu"],
            ["   - Darasa na Mzunguko (Safu E na F)"],
            [""],
            ["2. JAZA ALAMA ZA MASOMO (Safu G na kuendelea):"],
            ["MASOMO YOTE:"],
            [subjects_list],
            ["   - Alama kati ya 0 na 100"],
            ["   - Acha wazi kama hakushiriki"],
            ["   - Tumia namba tu"],
            [""],
            ["3. MAONI (Safu ya mwisho):"],
            ["   - Maoni ya jumla kwa mwanafunzi"],
            ["   - Si lazima kujaza"],
            [""],
            ["4. KWA WALIMU WA MASOMO:"],
            ["   - Kila mwalimu ajaze somo lake"],
            ["   - Mwalimu wa darasa achanganye"],
            [""],
            ["5. HIFADHI NA WASILISHA:"],
            [f"   - Jina la faili: {self._generate_filename()}"],
            ["   - Wasilisha kwa mwalimu mkuu"],
            ["   - Tarehe ya mwisho: [Weka tarehe]"],
            [""],
            ["⚠️ MUHIMU:"],
            ["- Usiongeze safu ziada"],
            ["- Usibadilishe mpangilio wa safu"],
            ["- Jinsia lazima iwe M (mvulana) au F (msichana)"],
            ["- Alama lazima ziwe kati ya 0 na 100"],
            [""],
            ["📞 Mawasiliano: [Mwalimu Mkuu/Msimamizi]"]
        ]
        
        # Write instructions
        for row_idx, instruction in enumerate(instructions, start=1):
            for col_idx, text in enumerate(instruction, start=1):
                cell = instructions_ws.cell(row=row_idx, column=col_idx, value=text)
                
                # Style based on content
                if row_idx == 1:  # Title
                    cell.font = Font(bold=True, color="366092", size=14)
                    cell.alignment = Alignment(horizontal="center")
                elif any(keyword in str(text) for keyword in ["TAARIFA", "JINSI YA", "MAONI", "KWA WALIMU", "HIFADHI", "MUHIMU"]):
                    cell.font = Font(bold=True, color="000000", size=11)
        
        # Adjust column width
        instructions_ws.column_dimensions['A'].width = 70
        
        # Add signature area
        sig_row = len(instructions) + 3
        instructions_ws.cell(row=sig_row, column=1, value="Sahihi ya Mwalimu wa Darasa: ________________________")
        instructions_ws.cell(row=sig_row, column=1).font = Font(italic=True)
        
        instructions_ws.cell(row=sig_row+1, column=1, value="Tarehe: ________________________")
        instructions_ws.cell(row=sig_row+1, column=1).font = Font(italic=True)
    
    def _generate_filename(self):
        """Generate filename for mark sheet"""
        class_name = self.class_info.get('name', 'Class').replace(' ', '_')
        stream = self.class_info.get('stream', '')
        
        if stream:
            return f"Full_MarkSheet_{class_name}_{stream}.xlsx"
        else:
            return f"Full_MarkSheet_{class_name}.xlsx"
    
    def get_template_summary(self):
        """Get summary of template"""
        return {
            'class': self.class_info.get('name', 'Form 4'),
            'stream': self.class_info.get('stream', ''),
            'filename': self._generate_filename(),
            'student_count': len(self.processed_data) if self.processed_data is not None else 0,
            'subjects': self.subjects,
            'total_columns': 6 + len(self.subjects) + 1,  # Student info + subjects + remarks
            'grading_rules': self.grading_rules,
            'instructions': 'Jaza alama kwenye safu za masomo tu. Usihariri taarifa za mwanafunzi.'
        }