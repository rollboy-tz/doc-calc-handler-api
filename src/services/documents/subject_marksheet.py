# services/documents/subject_marksheet.py
"""
SUBJECT MARKSHEET TEMPLATE - Tanzania NECTA
For subject teachers to fill marks for ONE subject only
No title rows - Clean headers only
"""
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from .base_document import BaseDocument

class SubjectMarkSheet(BaseDocument):
    """
    Single Subject Mark Sheet for subject teachers
    
    Features:
    - Student names pre-filled
    - ONE subject column only
    - Gender column included
    - Teacher fills marks for their subject only
    """
    
    def __init__(self, student_list, subject_info=None, **kwargs):
        """
        Initialize subject mark sheet
        
        Args:
            student_list: List of students
            subject_info: Dictionary with subject details
            **kwargs: Additional parameters
        """
        super().__init__(student_list, document_type='subject_marksheet', **kwargs)
        
        self.subject_info = subject_info or {}
        self.subject_name = subject_info.get('name', 'Mathematics')
        self.subject_code = subject_info.get('code', 'MATH')
        self.max_score = subject_info.get('max_score', 100)
        self.grading_rules = subject_info.get('grading_rules', 'CSEE')
        
        # Set metadata
        self.metadata.update({
            'subject_name': self.subject_name,
            'subject_code': self.subject_code,
            'max_score': self.max_score,
            'grading_rules': self.grading_rules,
            'student_count': len(student_list) if student_list else 0
        })
    
    def generate(self):
        """Generate the subject mark sheet template"""
        # Prepare student data with GENDER column
        students_df = self._prepare_student_data()
        
        # Add subject column (EMPTY for teacher to fill)
        students_df[self.subject_name] = ''  
        
        # Add remarks column
        students_df['Remarks'] = ''  
        
        self.processed_data = students_df
        return self
    
    def _prepare_student_data(self):
        """Prepare student data for template with GENDER column"""
        students = self.raw_data
        
        if not students:
            # Create sample data for testing
            students = [
                {
                    'admission_no': 'ADM001', 
                    'student_id': 'STU001', 
                    'full_name': 'JOHN MWAMBA', 
                    'gender': 'M',
                    'class': 'Form 4', 
                    'stream': 'East'
                },
                {
                    'admission_no': 'ADM002', 
                    'student_id': 'STU002', 
                    'full_name': 'SARAH JUMANNE', 
                    'gender': 'F',
                    'class': 'Form 4', 
                    'stream': 'East'
                }
            ]
        
        # Create DataFrame with required columns
        df = pd.DataFrame(students)
        
        # Ensure required columns exist (INCLUDING GENDER)
        required_columns = ['admission_no', 'student_id', 'full_name', 'gender', 'class', 'stream']
        
        for col in required_columns:
            if col not in df.columns:
                if col == 'gender':
                    df[col] = 'M'  # Default to Male
                elif col in ['class', 'stream']:
                    df[col] = 'NOT SET'
                else:
                    df[col] = f'{col.upper()}_MISSING'
        
        # Reorder columns - GENDER comes after full_name
        column_order = ['admission_no', 'student_id', 'full_name', 'gender', 'class', 'stream']
        df = df[column_order]
        
        return df
    
    def to_excel_bytes(self):
        """Generate Excel template as bytes - NO TITLE ROWS"""
        if self.processed_data is None:
            self.generate()
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main data sheet - NO startrow, headers are row 1
            self.processed_data.to_excel(
                writer, 
                sheet_name=f'{self.subject_code}_MARKS', 
                index=False
            )
            
            workbook = writer.book
            worksheet = writer.sheets[f'{self.subject_code}_MARKS']
            
            # Apply formatting
            self._apply_excel_formatting(workbook, worksheet)
            
            # Add instructions sheet
            self._add_instructions_sheet(workbook)
        
        output.seek(0)
        return output
    
    def _apply_excel_formatting(self, workbook, worksheet):
        """Apply formatting to Excel worksheet - NO TITLE ROW"""
        # ⭐⭐⭐ NO TITLE ROW - Headers start at row 1 ⭐⭐⭐
        
        # Header styling
        header_fill = PatternFill(
            start_color="366092",  # Professional blue
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Student info columns styling (read-only)
        info_fill = PatternFill(
            start_color="F2F2F2",  # Light gray
            end_color="F2F2F2",
            fill_type="solid"
        )
        
        # Gender column styling
        gender_fill = PatternFill(
            start_color="E2EFDA",  # Light green
            end_color="E2EFDA",
            fill_type="solid"
        )
        
        # Subject column styling (editable - highlight)
        subject_fill = PatternFill(
            start_color="FFF2CC",  # Light yellow - HIGHLIGHT FOR TEACHER
            end_color="FFF2CC",
            fill_type="solid"
        )
        
        # Remarks column styling
        remarks_fill = PatternFill(
            start_color="DDEBF7",  # Light blue
            end_color="DDEBF7",
            fill_type="solid"
        )
        
        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply to all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                
                # Header row (row 1)
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Data rows (row 2 onwards)
                elif cell.row > 1:
                    col_letter = get_column_letter(cell.column)
                    
                    # Student info columns (A-C) - read-only
                    if cell.column <= 3:  # Columns A-C (admission_no, student_id, full_name)
                        cell.fill = info_fill
                        cell.font = Font(color="000000", size=10)
                    
                    # Gender column (D) - read-only but colored differently
                    elif cell.column == 4:  # Column D (gender)
                        cell.fill = gender_fill
                        cell.font = Font(color="000000", size=10, bold=True)
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Class & Stream columns (E-F) - read-only
                    elif cell.column <= 6:  # Columns E-F (class, stream)
                        cell.fill = info_fill
                        cell.font = Font(color="000000", size=10)
                    
                    # Subject column (G) - EDITABLE - HIGHLIGHTED
                    elif cell.column == 7:  # Column G (subject marks)
                        cell.fill = subject_fill
                        cell.font = Font(color="000000", size=10, bold=True)
                        cell.alignment = Alignment(horizontal="center")
                        cell.number_format = '0'  # Integer format
                    
                    # Remarks column (H) - editable comments
                    elif cell.column == 8:  # Column H (remarks)
                        cell.fill = remarks_fill
                        cell.font = Font(color="000000", size=9, italic=True)
        
        # Set column widths
        column_widths = {
            'A': 15,  # Admission No
            'B': 12,  # Student ID
            'C': 25,  # Full Name
            'D': 10,  # Gender (M/F)
            'E': 10,  # Class
            'F': 10,  # Stream
            'G': 15,  # Subject Marks
            'H': 30   # Remarks
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Freeze header row
        worksheet.freeze_panes = "A2"
        
        # Add data validation for gender column
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Gender validation (M or F only)
        dv_gender = DataValidation(
            type="list",
            formula1='"M,F"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Gender",
            error="Please enter M or F only"
        )
        dv_gender.add(f'D2:D{worksheet.max_row}')
        worksheet.add_data_validation(dv_gender)
        
        # Marks validation (0-100)
        dv_marks = DataValidation(
            type="decimal",
            operator="between",
            formula1=0,
            formula2=self.max_score,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Marks",
            error=f"Marks must be between 0 and {self.max_score}"
        )
        dv_marks.add(f'G2:G{worksheet.max_row}')
        worksheet.add_data_validation(dv_marks)
    
    def _add_instructions_sheet(self, workbook):
        """Add instructions sheet for subject teacher"""
        instructions_ws = workbook.create_sheet(title="MAELEKEZO")
        
        instructions = [
            ["📘 MAELEKEZO YA KUJAZA FOMU"],
            [""],
            ["TAARIFA ZA SOMO:"],
            [f"Somo: {self.subject_name}"],
            [f"Kifupi: {self.subject_code}"],
            [f"Alama za juu: {self.max_score}"],
            [f"Mfumo wa upimaji: NECTA {self.grading_rules}"],
            [""],
            ["📝 JINSI YA KUJAZA:"],
            ["1. USIBADILISHE SEHEMU HII:"],
            ["   - Namba ya usajili (Safu A)"],
            ["   - Kitambulisho cha mwanafunzi (Safu B)"],
            ["   - Jina kamili (Safu C)"],
            ["   - Jinsia (Safu D) - M au F tu"],
            ["   - Darasa na Mzunguko (Safu E na F)"],
            [""],
            ["2. JAZA ALAMA ZA SOMO HILI PEKEE (Safu G):"],
            [f"   - Alama za {self.subject_name} pekee"],
            [f"   - Kati ya 0 na {self.max_score}"],
            ["   - Acha wazi kama hakushiriki"],
            ["   - Tumia namba tu (si alama za asilimia)"],
            [""],
            ["3. UNAWEZA KUONGEZA MAONI (Safu H):"],
            ["   - 'Hakushiriki' kama hakuja mtihani"],
            ["   - 'Haijakamilika' kama sehemu ya mtihani"],
            ["   - Maoni yoyote muhimu"],
            [""],
            ["4. KWA MWALIMU WA DARASA:"],
            ["   - Fomu hii ni kwa SOMO MOJA tu"],
            ["   - Wengine watajaza fomu zao"],
            ["   - Mwalimu wa darasa atachanganya zote"],
            [""],
            ["5. HIFADHI NA WASILISHA:"],
            [f"   - Jina la faili: {self._generate_filename()}"],
            ["   - Wasilisha kwa mwalimu wa darasa"],
            ["   - Tarehe ya mwisho: [Weka tarehe]"],
            [""],
            ["⚠️ MUHIMU:"],
            ["- Usiongeze safu ziada"],
            ["- Usibadilishe mpangilio wa safu"],
            ["- Wasilisha alama za somo hili tu"],
            ["- Jinsia lazima iwe M (kwa mvulana) au F (kwa msichana)"],
            [""],
            ["📞 Mawasiliano: [Mwalimu wa Darasa/Msimamizi]"]
        ]
        
        # Write instructions
        for row_idx, instruction in enumerate(instructions, start=1):
            for col_idx, text in enumerate(instruction, start=1):
                cell = instructions_ws.cell(row=row_idx, column=col_idx, value=text)
                
                # Style based on content
                if row_idx == 1:  # Title
                    cell.font = Font(bold=True, color="366092", size=14)
                    cell.alignment = Alignment(horizontal="center")
                elif any(keyword in str(text) for keyword in ["TAARIFA", "JINSI YA", "UNAWEZA", "KWA MWALIMU", "HIFADHI", "MUHIMU"]):
                    cell.font = Font(bold=True, color="000000", size=11)
                elif "⚠️" in str(text):
                    cell.font = Font(bold=True, color="FF0000", size=12)
                elif "📞" in str(text):
                    cell.font = Font(bold=True, color="00B050", size=11)
        
        # Adjust column width
        instructions_ws.column_dimensions['A'].width = 70
        
        # Add teacher signature area
        sig_row = len(instructions) + 3
        instructions_ws.cell(row=sig_row, column=1, value="Sahihi ya Mwalimu: ________________________")
        instructions_ws.cell(row=sig_row, column=1).font = Font(italic=True)
        
        instructions_ws.cell(row=sig_row+1, column=1, value="Tarehe: ________________________")
        instructions_ws.cell(row=sig_row+1, column=1).font = Font(italic=True)
    
    def _generate_filename(self):
        """Generate filename for subject mark sheet"""
        class_name = self.subject_info.get('class_name', 'Class').replace(' ', '_')
        stream = self.subject_info.get('stream', '')
        subject_code = self.subject_code
        term = self.subject_info.get('term', 'Term1')
        
        if stream:
            return f"MarkSheet_{subject_code}_{class_name}_{stream}.xlsx"
        else:
            return f"MarkSheet_{subject_code}_{class_name}.xlsx"
    
    def get_template_summary(self):
        """Get summary of template"""
        return {
            'subject': self.subject_name,
            'filename': self._generate_filename(),
            'student_count': len(self.processed_data) if self.processed_data is not None else 0,
            'columns': [
                'Admission No', 'Student ID', 'Full Name', 'Gender (M/F)', 
                'Class', 'Stream', f'{self.subject_name} Marks', 'Remarks'
            ],
            'max_score': self.max_score,
            'grading_rules': self.grading_rules,
            'instructions': f'Jaza alama za {self.subject_name} kwenye safu G tu. Usihariri taarifa za mwanafunzi.'
        }