"""
SINGLE SUBJECT TEMPLATE GENERATOR - PRODUCTION READY
For one subject only
"""
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .base_generator import BaseGenerator

class SubjectGenerator(BaseGenerator):
    """Generate single subject marksheet template"""
    
    def __init__(self, subject_name="MATHEMATICS", class_name="FORM 4", stream="", students=None):
        self.subject_name = subject_name
        self.class_name = class_name
        self.stream = stream
        self.students = students or []  # List of student dictionaries
    
    def generate(self):
        """Generate Excel template with actual student names"""
        if not self.students:
            return self._generate_sample_template()
        
        # Create DataFrame from actual students
        df = pd.DataFrame(self.students)
        
        # Ensure required columns exist
        required_columns = ['admission_no', 'student_id', 'full_name', 'gender']
        for col in required_columns:
            if col not in df.columns:
                df[col] = ''
        
        # Add class and stream columns if not present
        if 'class' not in df.columns:
            df['class'] = self.class_name
        if 'stream' not in df.columns:
            df['stream'] = self.stream
        
        # Add subject column (empty for teacher to fill)
        if self.subject_name not in df.columns:
            df[self.subject_name] = ''
        
        # Add remarks column if not present
        if 'remarks' not in df.columns:
            df['remarks'] = ''
        
        # Reorder columns
        final_columns = ['admission_no', 'student_id', 'full_name', 'gender', 'class', 'stream', self.subject_name, 'remarks']
        available_columns = [col for col in final_columns if col in df.columns]
        df = df[available_columns]
        
        return self._create_excel_file(df)
    
    def _generate_sample_template(self):
        """Generate template with sample data if no students provided"""
        # Create DataFrame structure
        data = {
            'admission_no': ['ADM001', 'ADM002', 'ADM003'],
            'student_id': ['STU001', 'STU002', 'STU003'],
            'full_name': ['JOHN DOE', 'JANE SMITH', 'MIKE JOHNSON'],
            'gender': ['M', 'F', 'M'],
            'class': [self.class_name, self.class_name, self.class_name],
            'stream': [self.stream, self.stream, self.stream],
            self.subject_name: ['', '', ''],
            'remarks': ['', '', '']
        }
        
        df = pd.DataFrame(data)
        return self._create_excel_file(df)
    
    def _create_excel_file(self, df):
        """Create Excel file from DataFrame"""
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = f"{self.subject_name}_MARKS"
            sheet_name = sheet_name[:31]  # Excel sheet name limit
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Apply formatting
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Basic formatting
            self._apply_basic_formatting(worksheet)
            
            # Highlight subject column (find its position)
            subject_col_index = None
            for idx, col in enumerate(df.columns, start=1):
                if col == self.subject_name:
                    subject_col_index = idx
                    break
            
            if subject_col_index:
                # Highlight subject column
                subject_fill = PatternFill(
                    start_color="FFF2CC",  # Light yellow
                    end_color="FFF2CC",
                    fill_type="solid"
                )
                
                subject_col_letter = get_column_letter(subject_col_index)
                for row in range(2, len(df) + 2):  # Skip header
                    cell = worksheet.cell(row=row, column=subject_col_index)
                    cell.fill = subject_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.number_format = '0'  # Integer format
            
            # Set column widths
            column_widths = {
                'A': 15,  # admission_no
                'B': 12,  # student_id
                'C': 25,  # full_name
                'D': 10,  # gender
                'E': 10,  # class
                'F': 10,  # stream
                'G': 15,  # subject marks (adjust if different position)
                'H': 25   # remarks
            }
            
            # Adjust for actual column positions
            for i, col_letter in enumerate(column_widths.keys(), start=1):
                if i <= len(df.columns):
                    worksheet.column_dimensions[col_letter].width = column_widths[col_letter]
            
            # Add instructions
            self._add_instructions_sheet(workbook)
        
        output.seek(0)
        return output
    
    def _add_instructions_sheet(self, workbook):
        """Add instructions sheet"""
        instructions_ws = workbook.create_sheet(title="INSTRUCTIONS")
        
        instructions = [
            [f"📘 {self.subject_name} MARK SHEET TEMPLATE"],
            [""],
            ["JINSI YA KUTUMIA:"],
            [f"1. Jaza alama za {self.subject_name} kwenye safu iliyopakwa rangi njano"],
            ["2. USIBADILI safu A-F (taarifa za mwanafunzi)"],
            ["3. Tumia namba pekee (0-100)"],
            ["4. Acha wazi kama mwanafunzi hayupo"],
            ["5. Weka maoni kwenye safu ya mwisho ikiwa inahitajika"],
            [""],
            ["HIFADHI NA PEEKISHA:"],
            ["1. Hifadhi faili iliyojazwa"],
            ["2. Peekisha kwa /api/extract/single-subject"],
            [f"3. Mfumo utachakua alama za {self.subject_name}"]
        ]
        
        for row_idx, instruction in enumerate(instructions, start=1):
            cell = instructions_ws.cell(row=row_idx, column=1, value=instruction[0])
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="366092")
    
    def get_info(self):
        """Get template information"""
        return {
            'type': 'single_subject',
            'subject': self.subject_name,
            'class': self.class_name,
            'stream': self.stream,
            'student_count': len(self.students) if self.students else 3,
            'columns': ['admission_no', 'student_id', 'full_name', 'gender', 'class', 'stream', self.subject_name, 'remarks'],
            'filename': f"{self.subject_name}_Marksheet_{self.class_name}.xlsx"
        }