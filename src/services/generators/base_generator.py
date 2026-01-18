"""
BASE GENERATOR - Common functionality for all generators
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from io import BytesIO

class BaseGenerator:
    """Base class for all template generators"""
    
    def _apply_basic_formatting(self, worksheet):
        """Apply basic formatting to worksheet"""
        # Header styling
        header_fill = PatternFill(
            start_color="366092",  # Professional blue
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply to header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Freeze header row
        worksheet.freeze_panes = "A2"
        
        # Add thin borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def _set_column_widths(self, worksheet, widths):
        """Set column widths"""
        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width