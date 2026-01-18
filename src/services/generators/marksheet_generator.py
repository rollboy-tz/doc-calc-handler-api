"""
FULL MARK SHEET TEMPLATE GENERATOR - PRODUCTION READY
For all subjects in one sheet
"""
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from .base_generator import BaseGenerator

class MarksheetGenerator(BaseGenerator):
    """Generate full marksheet template for all subjects"""
    
    def __init__(self, class_name="FORM 4", stream="", subjects=None, students=None):
        self.class_name = class_name
        self.stream = stream
        self.subjects = subjects or [
            "MATHEMATICS", "ENGLISH", "KISWAHILI",
            "PHYSICS", "CHEMISTRY", "BIOLOGY",
            "GEOGRAPHY", "HISTORY", "CIVICS", "COMMERCE"
        ]
        self.students = students or []  # List of student dictionaries
    
    def generate(self):
        """Generate Excel template with actual student names"""
        if not self.students:
            return self._generate_sample_template()
        
        # Create DataFrame from actual students
        df = pd.DataFrame(self.students)
        
        # Ensure required columns exist
        required_columns = ['admission_no', 'student_id', 'full_name', 'gender']
        for col in required_columns:
            if col not in df.columns:
                df[col] = ''
        
        # Add class and stream columns if not present
        if 'class' not in df.columns:
            df['class'] = self.class_name
        if 'stream' not in df.columns:
            df['stream'] = self.stream
        
        # Add empty subject columns
        for subject in self.subjects:
            if subject not in df.columns:
                df[subject] = ''
        
        # Add remarks column if not present
        if 'remarks' not in df.columns:
            df['remarks'] = ''
        
        # Reorder columns
        base_columns = ['admission_no', 'student_id', 'full_name', 'gender', 'class', 'stream']
        final_columns = base_columns + self.subjects + ['remarks']
        
        # Keep only columns that exist
        available_columns = [col for col in final_columns if col in df.columns]
        df = df[available_columns]
        
        return self._create_excel_file(df)
    
    def _generate_sample_template(self):
        """Generate template with sample data if no students provided"""
        # Create DataFrame structure
        data = {
            'admission_no': ['ADM001', 'ADM002', 'ADM003'],
            'student_id': ['STU001', 'STU002', 'STU003'],
            'full_name': ['JOHN DOE', 'JANE SMITH', 'MIKE JOHNSON'],
            'gender': ['M', 'F', 'M'],
            'class': [self.class_name, self.class_name, self.class_name],
            'stream': [self.stream, self.stream, self.stream]
        }
        
        # Add subject columns
        for subject in self.subjects:
            data[subject] = ['', '', '']
        
        # Add remarks
        data['remarks'] = ['', '', '']
        
        df = pd.DataFrame(data)
        return self._create_excel_file(df)
    
    def _create_excel_file(self, df):
        """Create Excel file from DataFrame"""
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = f"{self.class_name}_{self.stream}" if self.stream else self.class_name
            sheet_name = sheet_name[:31]  # Excel sheet name limit
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Apply formatting
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Basic formatting
            self._apply_basic_formatting(worksheet)
            
            # Set column widths
            column_widths = {
                'A': 15,  # admission_no
                'B': 12,  # student_id
                'C': 25,  # full_name
                'D': 10,  # gender
                'E': 10,  # class
                'F': 10,  # stream
            }
            
            # Subject columns (12 width each)
            subject_start_col = 7  # Column G
            for i, subject in enumerate(self.subjects):
                col_letter = get_column_letter(subject_start_col + i)
                column_widths[col_letter] = 15
            
            # Remarks column
            remarks_col = get_column_letter(subject_start_col + len(self.subjects))
            column_widths[remarks_col] = 25
            
            self._set_column_widths(worksheet, column_widths)
            
            # Add instructions sheet
            self._add_instructions_sheet(workbook)
        
        output.seek(0)
        return output
    
    def _add_instructions_sheet(self, workbook):
        """Add instructions sheet"""
        instructions_ws = workbook.create_sheet(title="INSTRUCTIONS")
        
        instructions = [
            ["📘 FULL MARK SHEET TEMPLATE"],
            [""],
            ["JINSI YA KUTUMIA:"],
            ["1. USIBADILI safu A-F (taarifa za mwanafunzi)"],
            ["2. Jaza alama kwenye safu za masomo (G na kuendelea)"],
            ["3. Tumia namba pekee (0-100)"],
            ["4. Acha wazi kama mwanafunzi hayupo"],
            ["5. Weka maoni kwenye safu ya mwisho ikiwa inahitajika"],
            [""],
            ["HIFADHI NA PEEKISHA:"],
            ["1. Hifadhi faili iliyojazwa"],
            ["2. Peekisha kwa /api/extract/multi-subject"],
            ["3. Mfumo utachukua na kuchakua data"]
        ]
        
        for row_idx, instruction in enumerate(instructions, start=1):
            cell = instructions_ws.cell(row=row_idx, column=1, value=instruction[0])
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="366092")
    
    def get_info(self):
        """Get template information"""
        return {
            'type': 'full_marksheet',
            'class': self.class_name,
            'stream': self.stream,
            'subjects': self.subjects,
            'student_count': len(self.students) if self.students else 3,
            'student_columns': ['admission_no', 'student_id', 'full_name', 'gender', 'class', 'stream'],
            'total_columns': 6 + len(self.subjects) + 1,
            'filename': f"Marksheet_{self.class_name}_{self.stream}.xlsx" if self.stream else f"Marksheet_{self.class_name}.xlsx"
        }